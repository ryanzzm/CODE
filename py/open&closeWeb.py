#!/usr/bin/python
import time
import webbrowser
import os
import datetime
from openpyxl import load_workbook

#first run...
#check time to locate next full hour
currentDateTime = datetime.datetime.now() #full date and time
currentYear = currentDateTime.year
currentMonth = currentDateTime.month
currentDay = currentDateTime.day
currentHour = currentDateTime.hour
currentMin = currentDateTime.minute
currentSec = currentDateTime.second

#time.sleep(3600-currentMin*60-currentSec)
waitTime = 60-currentSec
print("wait: %d" % waitTime)
time.sleep(waitTime)


i = 1
startRow = 1
while i == 1 :
    currentDateTime = datetime.datetime.now() #full date and time
    currentYear = currentDateTime.year
    currentMonth = currentDateTime.month
    currentDay = currentDateTime.day
    currentHour = currentDateTime.hour
    currentMin = currentDateTime.minute
    currentSec = currentDateTime.second
    currentWeekNumber = currentDateTime.isocalendar()[1] # week number of the year
    currentWeekdayNumber = currentDateTime.isocalendar()[2] # weekday number of the week
    #open and check spreadsheet
    wb = load_workbook(filename='youtube.xlsm', read_only=False)
    ws = wb['play_list']
    maxRowNumber = ws.max_row
    if currentHour > 10 and currentHour < 16 and currentWeekdayNumber < 6 :
    #if currentMin > 0 and currentMin < 60 and currentWeekdayNumber < 6 :
        print(startRow)
        print(maxRowNumber)
        if startRow < maxRowNumber:
            pageURL = ws.cell(row=startRow+1, column=1).value
        else:
            pageURL = ws.cell(row=2, column=1).value
            startRow = 1
        startRow = startRow + 1
        print(pageURL)
        webbrowser.open_new_tab(pageURL)  # Go to example.com
        time.sleep(120)
        os.system("TASKKILL /F /IM chrome.exe")
        time.sleep(3480)
    else:
        time.sleep(3600)
        print("out")
       
